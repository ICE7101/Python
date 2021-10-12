from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill, Font
import os

print('''把总表那个EXCEL取名为total，里面的sheet名为Details。
总表里分行号一定要在G列。总表和分行EXCEL的客户号都在A列。分行的EXCEL里deposit_gbp的值一定要在H列。
分行的EXCEL文件名应该含有分行号，例如BD。
然后把所有EXCEL放入这个exe文件同一文件夹。
*****************************************
''')
lastMonth=input('输入上个月份数据所在的sheet名：')
print('''*****************************************
''')
ThisMonth=input('输入本月份作为新sheet的名字，例如202012。')

filenames1=os.listdir('./')
BranchList=['BD','WE','BB','GB','MB']
if lastMonth[-2:]=='01':
    month='Jan'
elif lastMonth[-2:]=='02':
    month='Feb'
elif lastMonth[-2:]=='03':
    month='Mar'
elif lastMonth[-2:]=='04':
    month='Apr'
elif lastMonth[-2:]=='05':
    month='May'
elif lastMonth[-2:]=='06':
    month='Jun'
elif lastMonth[-2:]=='07':
    month='Jul'
elif lastMonth[-2:]=='08':
    month='Aug'
elif lastMonth[-2:]=='09':
    month='Sep'
elif lastMonth[-2:]=='10':
    month='Oct'
elif lastMonth[-2:]=='11':
    month='Nov'
elif lastMonth[-2:]=='12':
    month='Dec'
#把文件夹里的文件名去匹配各分行名称。含有分行名称就打开total表
for filename in filenames1:
    for branch in BranchList:
        if branch in filename:
#*********打开含有branch名的表
            wbbranch = load_workbook('./'+filename)
            wsbranch = wbbranch[lastMonth]
            # 新建一个新的sheet
            wbbranch.create_sheet(ThisMonth) 
            wsbranch_new=wbbranch[ThisMonth]
        # ****************打开工作簿总表，获取活动工作表
            wb = load_workbook('./total.xlsx')
            ws = wb['Details']
            #给新sheet加上列名那一行
            for row in ws.iter_rows(max_row=1, values_only=True):
                wsbranch_new.append(row[:8]+('deposit_gbp '+month,'Difference')+row[8:])
            for row in ws.iter_rows(min_row=2, values_only=True):
                #总表这行的Team列的值等于分行名，才处理     
                if row[6]==branch:  
                    
                    #把主表这行的客户号放到branch表里去循环匹配
                    for branch_row in wsbranch.iter_rows(min_row=2, values_only=True): 
                    #不同excel里客户号列可能一个是数字型，一个是字符型，所以都转为字符型来比较         
                        if str(branch_row[0])==str(row[0]):
                            difference=float(row[7])-float(branch_row[7])
                            branch_lastMonth=branch_row[7]
                            break
                        else:
                            difference=0
                            branch_lastMonth=''
                    wsbranch_new.append(row[:8]+ (branch_lastMonth,difference)+row[8:])
                            
            wsbranch_new['I1'].fill=PatternFill(patternType='solid', fgColor='FFFF00')
            wsbranch_new['J1'].fill=PatternFill(patternType='solid', fgColor='FFCC00')
            for i in range(ord('A'), ord('S')+1):
                if chr(i)!='I' and chr(i)!='J':
                    x=chr(i)+'1'
                    wsbranch_new[x].fill=PatternFill(patternType='solid', fgColor='001166')
                    wsbranch_new[x].font=Font(color='ffffff',bold=True)
            for row in wsbranch_new.iter_rows(min_row=2):
                for i in range (7,21):
                
                    row[i].number_format ='#,##0'        
            


            wbbranch.save('./'+filename)